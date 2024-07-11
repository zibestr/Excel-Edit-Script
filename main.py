import os
import re
from dataclasses import dataclass

from dadata import Dadata
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

# загружаем переменную среду
load_dotenv('.env')

token = os.getenv('API_TOKEN')  # токен API
dadata = Dadata(token)  # доступ к API Dadata
colleges = set()


class NonActiveOrganizationException(Exception):
    """
    Вызывается если по заданному ИНН ликвидированы все организации
    """
    pass


@dataclass
class Organization:
    """
    Класс, предоставляющий доступ к информации об организации.

    При инициализации подается ИНН и КПП (КПП нужен для определения
    конкретного филиала ВУЗа, если по данному ИНН зарегистрированно
    больше одной организации).

    Если все организации ликвидированы по данному ИНН, то вызывает исключение
    NonActiveOrganizationException
    """
    _data = None

    def __init__(self, row):
        party = row[0].value
        kpp = row[2].value
        if len(row) == 6:
            self._short_name = row[5].value
        else:
            self._short_name = row[4].value
        # в таблице могут податься значения float с точкой, вместо str
        party = float2str(party)

        response = dadata.find_by_id('party', party)
        if not response and len(party) < 10:
            party = '0' + party
            response = dadata.find_by_id('party', party)
        self.inn = party

        # если по заданному ИНН зарегистрированы еще и филиалы
        if len(response) > 1:
            if kpp:
                kpp = float2str(kpp)
                if len(kpp) < 9:
                    kpp = '0' + kpp

                # сверяем КПП из таблицы с КПП из ответа API
                for i, org in enumerate(response):
                    if org['data']['kpp'] == kpp:
                        # Проверка на активность организации
                        if org['data']['state']['status'] == "ACTIVE":
                            self._data = org['data']
                    # Если по заданному ИНН и КПП не найдена действующая
                    # организация, то вызывает ошибку
                    if i == len(response) - 1 and self._data is None:
                        raise NonActiveOrganizationException()
        else:
            self._data = response[0]['data']

        # Если найденная организация ликвидирована, то вызывает исключение
        if self._data['state']['status'] != "ACTIVE":
            raise NonActiveOrganizationException()

        if (self.inn, self.ogrn, self.kpp) in colleges:
            if len(row) == 6:
                write_copy_organization(row[4].value)
            else:
                write_copy_organization(row[3].value)
            raise ValueError('ВУЗ уже добавлен в таблицу')
        else:
            colleges.add((self.inn, self.ogrn, self.kpp))

    @staticmethod
    def edit_full_name(full_name: str) -> str:
        full_name = full_name.lower().capitalize()
        if full_name.count('"') > 2:
            full_name = re.sub(r'\"(.*\".*)\"',
                               r'«\g<1>»',
                               full_name).replace('"', '')
        else:
            full_name = re.sub(r'"(.+?)"', r'«\1»', full_name)
        new_full_name = ''
        is_start = False
        for i in range(len(full_name)):
            if full_name[i] == '«':
                is_start = True
            elif full_name[i] == '»':
                is_start = False
            if full_name[i - 1] == '«' \
               or is_start and full_name[i - 1] == ' ' \
               or full_name[i - 7: i] == 'городе ' \
               or full_name[i - 1] == '.' \
               or full_name[i - 2:i] == '. ' \
               or full_name[i - 2:i] == 'г ':
                new_full_name += full_name[i].capitalize()
            else:
                new_full_name += full_name[i]
            if full_name[i] == ';':
                break
        return new_full_name

    # TODO: редактирование названия
    @property
    def full_name(self) -> str:
        return self.edit_full_name(self._data['name']['full_with_opf'])

    @property
    def short_name(self) -> str | None:
        if self._short_name is None or self._short_name == '-':
            if self._data['name']['short'] is None:
                return None
            return self._data['name']['short'].split(',')[0]
        return self._short_name

    @property
    def ogrn(self) -> str:
        return self._data['ogrn']

    @property
    def kpp(self) -> str:
        return self._data['kpp']


def float2str(value: float | str) -> str:
    if isinstance(value, float):
        value = str(value).split('.')[0].strip()
    else:
        value = value.strip()
    return value


# если найдена ликвидированная организация - записываем ее в файл
def write_non_active_organization(inn, ogrn, kpp):
    with open('Ликвидированные организации.txt', 'a') as file:
        file.write(f'ИНН {inn}, '
                   f'ОГРН {ogrn}, КПП {kpp}\n')


# если вуз уже записан в таблицу - записываем в файл
def write_copy_organization(full_name: str):
    with open('Повторяющие вузы.txt', 'a') as file:
        file.write(full_name + '\n')


def validate_sheet(sheet, new_sheet):
    if sheet.title == 'База 6':
        max_col = 7
    else:
        max_col = 6

    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=max_col):
        # прерываем поиск по листу, если лист закончился
        if row[0].value is None:
            break

        try:
            org = Organization(row)
        except NonActiveOrganizationException:
            inn, ogrn, kpp = float2str(row[0].value), \
                float2str(row[1].value), float2str(row[2].value)
            print(f'Найдена ликвидированная организация: ИНН {inn}, '
                  f'ОГРН {ogrn}, КПП {kpp}')
            write_non_active_organization(inn, ogrn, kpp)
            continue
        except ValueError:
            continue

        new_sheet.append([org.inn, org.ogrn,
                          org.kpp, org.full_name,
                          org.short_name])

    return new_sheet


if __name__ == '__main__':
    # очищаем файлы
    with open('Ликвидированные организации.txt', 'w') as file:
        file.write('')
    with open('Повторяющие вузы.txt', 'w') as file:
        file.write('')

    # загрузка оригинальной таблицы
    table = load_workbook('data/Справочник вузов.xlsx')
    table.remove(table['Правила'])
    table.remove(table['Лист5'])

    # создание новой таблицы
    new_table = Workbook()
    new_sheet = new_table.create_sheet('Общая База')
    new_sheet.append(['Inn', 'Ogrn',
                      'КПП', 'Полное наименование',
                      'Короткое наименование'])
    new_table.remove(new_table['Sheet'])

    # проходимся по каждому листу исходной таблицы
    for sheet in table:
        new_sheet = validate_sheet(sheet, new_sheet)
        print(f'Лист {sheet.title} обработан')

    # сохранение таблицы
    new_table.save('Справочник вузов (Редактированный).xlsx')
